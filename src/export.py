"""통제 설계안 엑셀(감사조서형) 내보내기.

시트 구성(실무자 피드백 반영):
  ① 전제조건(ITGC)  ② 경영진주장 리스크맵  ③ 통제설계안  ④ 사람검토 필요
'사람검토 필요' 카드 중 사용자가 검토·확정한 통제는 ③에 '검토 확정' 표시로 포함된다.
"""
from __future__ import annotations

import datetime
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

NAVY = "1F2A44"
ORANGE = "DB4E18"
GREY_FILL = "F2F4F8"
F = "Malgun Gothic"

_thin = Border(*[Side(style="thin", color="D9D9D9")] * 4)
_wrap = Alignment(wrap_text=True, vertical="top")


def _header(ws, row: int, headers: list[str], fill: str = NAVY) -> None:
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c, h)
        cell.font = Font(name=F, bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = _wrap
        cell.border = _thin


def _cell(ws, r: int, c: int, v, bold: bool = False, color: str = "000000"):
    cell = ws.cell(r, c, v)
    cell.font = Font(name=F, size=10, bold=bold, color=color)
    cell.alignment = _wrap
    cell.border = _thin
    return cell


def _widths(ws, widths: list[int]) -> None:
    for col, w in zip("ABCDEFGHIJK", widths):
        ws.column_dimensions[col].width = w


def _card_row(ws, r: int, card, origin: str) -> None:
    ev_txt = "\n".join(
        f"{e.source} — {e.paragraph} [{'원문 대조 완료' if e.verified else '원문 대조 전'}]"
        for e in card.evidences
    )
    _cell(ws, r, 1, card.control_id)
    _cell(ws, r, 2, f"층 {card.layer}")
    _cell(ws, r, 3, getattr(card, "activity_type", None) or "-")
    _cell(ws, r, 4, card.name, bold=True)
    _cell(ws, r, 5, card.why)
    _cell(ws, r, 6, card.implementation)
    _cell(ws, r, 7, ev_txt)
    _cell(ws, r, 8, card.confidence)
    _cell(ws, r, 9, "Y" if card.forced else "")
    _cell(ws, r, 10, origin,
          color=ORANGE if origin.startswith("사람 검토") else "000000")


def build_excel(output, ctx: dict, itgc_checks: list[dict],
                promoted_ids: set[str]) -> bytes:
    """AgentOutput → xlsx bytes. promoted_ids: 사람검토에서 확정된 통제 id."""
    wb = Workbook()

    # ① 전제조건
    ws = wb.active
    ws.title = "전제조건(ITGC)"
    ws.cell(1, 1, "AIGC Architect — 통제 설계안 (설계 권고이며 판정이 아님)").font = \
        Font(name=F, size=13, bold=True, color=NAVY)
    meta = [("생성일시", datetime.datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("산업", ctx.get("industry", "")), ("기업 규모", ctx.get("size", "")),
            ("모델 조달", ctx.get("sourcing", "")),
            ("게이트 판정", "통과" if output.gate.passed else "미비(조건부 진행)")]
    r = 3
    for k, v in meta:
        _cell(ws, r, 1, k, bold=True)
        _cell(ws, r, 2, v)
        r += 1
    r += 1
    _header(ws, r, ["카테고리", "확인 질문", "판정", "왜 선결 조건인가"])
    for chk in itgc_checks:
        r += 1
        ok = chk["id"] not in output.gate.failed_checks
        _cell(ws, r, 1, chk.get("category", "기타"))
        _cell(ws, r, 2, chk["question"])
        _cell(ws, r, 3, "충족" if ok else "미비", bold=True,
              color="1A7F37" if ok else "B42318")
        _cell(ws, r, 4, chk["why_prerequisite"])
    _widths(ws, [30, 52, 8, 52])
    ws.freeze_panes = "A9"

    # ② 리스크맵
    ws = wb.create_sheet("경영진주장 리스크맵")
    _header(ws, 1, ["프로세스", "리스크", "영향받는 경영진주장"], ORANGE)
    r = 1
    for risk in output.risks:
        r += 1
        _cell(ws, r, 1, risk.process_name)
        _cell(ws, r, 2, risk.risk)
        _cell(ws, r, 3, risk.assertion_impact)
    _widths(ws, [24, 60, 40])
    ws.freeze_panes = "A2"

    # ③ 통제설계안 (수용 + 검토 확정)
    ws = wb.create_sheet("통제설계안")
    headers = ["통제ID", "층", "유형(보론3-20)", "통제명", "왜 필요한가",
               "구체적 실행 모습", "근거 조항(검증상태)", "신뢰도", "강제포함", "수록 경로"]
    _header(ws, 1, headers)
    r = 1
    ordered = sorted(output.harness.accepted, key=lambda c: (c.layer, c.control_id))
    for card in ordered:
        r += 1
        _card_row(ws, r, card, "자동 수용")
    for card in output.harness.quarantined:
        if card.control_id in promoted_ids:
            r += 1
            _card_row(ws, r, card, "사람 검토 후 확정")
    _widths(ws, [10, 6, 14, 24, 40, 48, 44, 8, 8, 14])
    ws.freeze_panes = "A2"

    # ④ 사람검토 필요 (미확정 잔여)
    ws = wb.create_sheet("사람검토 필요")
    _header(ws, 1, ["통제ID", "층", "유형(보론3-20)", "통제명", "왜 필요한가",
                    "구체적 실행 모습", "근거 조항(검증상태)", "신뢰도", "강제포함", "격리 사유"],
            "6B7280")
    r = 1
    for card in output.harness.quarantined:
        if card.control_id not in promoted_ids:
            r += 1
            _card_row(ws, r, card, f"신뢰도 {card.confidence} < 임계 0.60")
    _widths(ws, [10, 6, 14, 24, 40, 48, 44, 8, 8, 22])
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
